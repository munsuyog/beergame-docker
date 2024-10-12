# views.py
import os
import time
import json
import pickle
import atexit
import numpy
from math import floor, inf
from beergame.json_pprint import MyEncoder
from beergame import game, app
from flask import request, session, redirect, url_for, render_template, flash, jsonify, Response, send_from_directory
import logging
from logging.handlers import RotatingFileHandler
# from beergame.firebase_config import verify_id_token, db
from flask_cors import CORS, cross_origin
import math
from beergame.session import Session
from beergame.stations import Demand
import uuid
import datetime
import signal
import sys
import beergame.game
from beergame.mongo_config import db
from bson.objectid import ObjectId
import pymongo

################################################################################
# Logging setup
################################################################################
# logging.basicConfig(level=logging.DEBUG,format='%(name)-12s|%(message)s')
# file_log = RotatingFileHandler(filename='./instance/server.log', maxBytes=1048576, backupCount=9)
# file_log.setLevel(logging.DEBUG)
# file_log.setFormatter(logging.Formatter('%(asctime)s|%(name)-12s|%(message)s'))
# logging.getLogger().addHandler(file_log)
logger = logging.getLogger(__name__)


def log_msg(msg: str):
    logger.warning(msg)


def show_error(m: str, e: Exception):
    template = "   Error: {0}. Exception of type {1} occurred. Arguments:{2!r}."
    message = template.format(m, type(e).__name__, e.args)
    log_msg(message)
    flash(message)


################################################################################
# flask setup
################################################################################
log_msg('****** server started')
log_msg('   Local working folder location: ' + os.getcwd())
# log_msg('   Current instance location:     ' + app.instance_path)
app.config.from_object(__name__)  # load config from this file
try:
    with open('./beergame/config.json') as config_file:  # Load default config and override config from an environment variable
        config_data = json.load(config_file)
    log_msg('   Read config from ' + os.getcwd() + '/xsg/config.json')
except Exception as e:
    show_error('Error in reading XSG\'s config file',e)
# config_data['GameStatusData'] = os.path.join(app.instance_path, config_data['GameStatusData'])
config_data['GameStatusData'] = os.getcwd() + '/instance/' + config_data['GameStatusData']
app.secret_key = config_data['SECRET_KEY']  # set the secret key for 'session'
app.config.update(config_data)
app.config.from_envvar('XSG_SETTINGS', silent=True)
SecondsAway_to_Disconnect = int(config_data['SECONDSAWAY_TO_DISCONNECT'])
RefreshInterval_PlayScreen = int(config_data['SECONDS_TO_REFRESH_PLAY_SCREEN']) * 1000
RefreshInterval_GameMonitor = int(config_data['SECONDS_TO_REFRESH_GAME_MONITOR']) * 1000
RefreshInterval_GameDebug = int(config_data['SECONDS_TO_REFRESH_GAME_DEBUG']) * 1000
MaxGamesMonitored = int(config_data['MAX_GAMES_MONITORED'])
ALLOWED_EXTENSIONS = set(['json'])

CORS(app)

GAMES = {}
SESSIONS = {}

################################################################################
# output formating & utility functions
################################################################################
def handle_shutdown():
    if save_games_state():
        log_msg('Saved games state. Server shutting down...')
    else:
        log_msg('Failed to save games state. Server shutting down...')

atexit.register(handle_shutdown)



def kill_expired_games():
    d = []
    for k,v in GAMES.items():
        if (time.time() - v.created)/86400 > v.expiry:  # 86400 second in a day
            d.append(k)
    if len(d) > 0:
        log_msg("The following game(s) expired and will be deleted: {:}.".format(str(d)))
        for k in d:
            del GAMES[k]

def save_games_state():
    logger.info("Saving games state using pickle")
    try:
        kill_expired_games()
        with open(app.config['GameStatusData'], 'wb') as f:
            pickle.dump(GAMES, f)
        logger.info(f"Saved {len(GAMES)} games to {app.config['GameStatusData']}")
        return True
    except Exception as e:
        logger.exception(f"Error saving games state: {str(e)}")
        return False

def load_games_state():
    logger.info("Loading games state using pickle")
    try:
        if os.path.isfile(app.config['GameStatusData']):
            with open(app.config['GameStatusData'], 'rb') as f:
                GAMES.update(pickle.load(f))
            kill_expired_games()
            logger.info(f"Loaded previous game state from {app.config['GameStatusData']}")
            return True
        else:
            logger.warning(f"Previous game state file not found: {app.config['GameStatusData']}")
            return False
    except Exception as e:
        logger.exception(f"Error loading games state: {str(e)}")
        return False



# loading previous server game state data if existing
load_games_state()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def list_avg(x):
    if len(x) == 0:
        return 0
    else:
        return sum(x)/len(x)


def list_cum_avg(x):
    if len(x) == 0:
        return 0
    else:
        cum_sum = [numpy.asscalar(i) for i in numpy.cumsum(x)]
        cum_avg = [cum_sum[i]/(i+1) for i in range(len(cum_sum))]
        return cum_avg


def toHTMLtbl(x):
    if type(x) is list:
        return ' ' + str(x).replace('\'','').replace('[','').replace(']','').replace('(','<tr><td>')\
            .replace('), ','</td></tr>').replace(', ','</td><td>').replace(')','</td></tr>')


def txt2list(l):
    if isinstance(l, list):
        return l  # If it's already a list, return it as is
    try:
        return [float('inf') if x.strip().lower() == 'infinity' else int(x) for x in l.split(',')]
    except Exception as e:
        msg = f"Could not convert text list '{l}' to numeric list. Exception {type(e).__name__} - {e.args}."
        log_msg(msg)
        raise ValueError(msg)


def show_week(D,week):
    """show a given week's values as a dict"""
    r = {}
    for k,v in D.items():
        r[k] = v[week]
    return r


def week_sum(D,week):
    """ return a week's sum from a dictionary of lists of weekly values """
    return sum([x[week] for x in D.values()])

################################################################################
# application views
################################################################################
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')


@app.route('/robots.txt')
def robots():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'robots.txt', mimetype='text/plain')


@app.route('/cmd/<cmd>')
def cmd(cmd):
    if cmd == 'index':
        return redirect(url_for('index'))
    if cmd == 'about':
        return render_template('about.html')
    if cmd == 'logout':
        session.pop('username',None)
        session.pop('password',None)
        return redirect(url_for('cmd',cmd='reset'))
    if cmd == 'reset':
        session.pop('player_name',None)
        session.pop('selected_game',None)
        session.pop('selected_station',None)
        return redirect(url_for('index'))
    if cmd == 'import_game':
        return render_template('import_game.html')
    if cmd == 'advanced_menu':
        return render_template('advanced_menu.html')
    if cmd == 'admin_game_start':
        return select_game_page(text='Select a game, and fill in the game\'s admin password:',next_page='admin_game', ask_name=False, ask_password=True)
    if cmd == 'join_game':
        return select_game_page(text='Select the game you like to join, and fill in your info:',next_page='join_station', ask_name=True, ask_password=True)
    if cmd == 'create_game_from_template':
        templates = [x.rsplit('.',1)[0] for x in os.listdir('./game_templates/')]
        return render_template('create_game_from_template.html', data=templates)
    if cmd == 'monitor_games':
        kill_expired_games()
        return render_template('multi-game-select.html', MaxGamesMonitored=[x for x in range(MaxGamesMonitored)], games_list=[' ']+[x.team_name for x in GAMES.values()])
    if cmd == 'under_construction':
        return 'Sorry, this part is still under construction'
    return "Unknown cmd: " + cmd


# direct entry point, not reachable through WUI
@app.route('/admin')
def admin_server_start():
    if 'password' in session.keys() and 'username' in session.keys():
        if app.config['USERNAME'] == session['username'] and app.config['PASSWORD'] == session['password']:
            return redirect(url_for('admin_server_menu'))
    return render_template('admin_password.html')


@app.route('/admin_server_menu', methods=['POST','GET'])
def admin_server_menu():
    if request.method == 'GET':
        if 'password' in session.keys() and 'username' in session.keys():
            if app.config['USERNAME'] != session['username']:
                flash('User is not an admin!')
                return redirect(url_for('admin_server_start'))
            if app.config['PASSWORD'] != session['password']:
                flash('Incorrect password!')
                return redirect(url_for('admin_server_start'))
        else:
            return redirect(url_for('admin_server_start'))
    elif request.method == 'POST':
        if app.config['USERNAME'] != request.form['username']:
            flash('User is not an admin!')
            return redirect(url_for('admin_server_start'))
        if app.config['PASSWORD'] != request.form['password']:
            flash('Incorrect password!')
            return redirect(url_for('admin_server_start'))
        session['username'] = request.form['username']
        session['password'] = request.form['password']
    return render_template('admin_server_menu.html')


@app.route('/demo', methods=['GET'])
@cross_origin()
def demo():
    demo_filename = './game_templates/demo.json'
    kill_expired_games()
    import random
    id = '{:03d}'.format(random.randint(1, 999))
    while 'demo_'+id in GAMES.keys():
        id = '{:03d}'.format(random.randint(1, 999))
    if os.path.isfile(demo_filename):
        try:
            game_data = {}
            with open(demo_filename) as game_file:
                game_data = json.load(game_file)
            game_data['team_name'] = 'demo_'+id
            game_data['play_password'] = id
            game_data['admin_password'] = id
            GAMES[game_data['team_name']] = game.Game(game_data)
            session['player_name'] = 'DemoPlayer'
            session['selected_game'] = game_data['team_name']
            session['selected_station'] = 'Manufacturer'
            return jsonify({
                'id': id,
                'team_name': game_data['team_name'],
                'play_password': game_data['play_password'],
                'admin_password': game_data['admin_password'],
                'player_name': 'DemoPlayer',
                'selected_station': 'Manufacturer'
            })
        except Exception as e:
            if game_data.get('team_name','') in GAMES.keys():
                del GAMES[game_data['team_name']]
            show_error("Could not create demo game", e)
            return jsonify({'error': 'Could not create demo game'}), 500
    else:
        flash('Demo game JSON file not found.')
        return jsonify({'error': 'Demo game JSON file not found'}), 404


@app.route('/admin_game', methods=['POST','GET'])
def admin_game():
    if request.method == 'GET':
        if 'selected_game' in session.keys():
            this_game = session['selected_game']
        else:
            this_game = ''
        if this_game not in GAMES.keys():
            flash('Previous game is no longer valid. Please select another game.')
            return redirect(url_for('cmd',cmd='admin_game_start'))
    if request.method == 'POST':
        this_game = request.form.get('selected_game')
        if this_game == '** No games created yet! **':
            return redirect(url_for('index'))
        if this_game not in GAMES.keys():
            flash('Selected game is no longer valid. Please select another game.')
            return redirect(url_for('cmd',cmd='admin_game_start'))
        if GAMES[this_game].admin_password != 'admin':
            flash('Incorrect password!')
            return redirect(url_for('cmd',cmd='admin_game_start'))
        session['selected_game'] = this_game
    return render_template('admin_game.html', game=this_game)


# @app.route('/join_station', methods=['POST'])
# @cross_origin()
# def join_station():
#     try:
#         data = request.json
#         selected_game = data.get('selected_game')
#         player_name = data.get('player_name')
#         password = data.get('password')

#         if not selected_game or not player_name or not password:
#             return jsonify({'error': 'Missing required parameters'}), 400

#         if selected_game not in GAMES:
#             return jsonify({'error': 'Selected game is no longer valid. Please select another game.'}), 400

#         if GAMES[selected_game].play_password != password:
#             return jsonify({'error': 'Incorrect password!'}), 400

#         session['player_name'] = player_name
#         session['selected_game'] = selected_game

#         # Update the player name in the game and MongoDB
#         game = GAMES[selected_game]
#         updated_station = None
#         for station in game.network_stations.values():
#             if station.player_name.startswith('player') and station.player_name != player_name:
#                 station.player_name = player_name
#                 station.touch()
#                 updated_station = station
#                 break

#         if updated_station:
#             # Update MongoDB
#             result = db.games.update_one(
#                 {'_id': selected_game, 'stations.name': updated_station.station_name},
#                 {'$set': {
#                     'stations.$.player_name': player_name,
#                     'stations.$.last_communication_time': updated_station.last_communication_time
#                 }}
#             )
#             if result.modified_count == 0:
#                 logger.warning(f"Failed to update player name in MongoDB for game {selected_game}, station {updated_station.station_name}")
#         else:
#             logger.warning(f"No available station found for player {player_name} in game {selected_game}")
#             return jsonify({'error': 'No available station found'}), 400

#         if game.manual_stations_names:
#             stations = game.manual_stations_names
#             return jsonify({'stations': stations, 'assigned_station': updated_station.station_name}), 200
#         else:
#             if game.Run():
#                 return jsonify({'message': 'Game simulation complete. (All stations are fully automated, see game config)'}), 200
#             else:
#                 return jsonify({'message': 'Game already simulated. Nothing done.'}), 200
#     except pymongo.errors.PyMongoError as e:
#         logger.exception(f"A MongoDB error occurred in join_station: {str(e)}")
#         return jsonify({'error': f'A database error occurred: {str(e)}'}), 500
#     except Exception as e:
#         logger.exception(f"An error occurred in join_station: {str(e)}")
#         return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    

# @app.route('/play_screen', methods=['GET'])
# @cross_origin()
# def play_screen():
#     this_game = request.args.get('selected_game')
#     this_station = request.args.get('selected_station')

#     if not this_game or not this_station:
#         return jsonify({'error': 'No game or station selected'}), 400

#     if this_game not in GAMES.keys():
#         return jsonify({'error': 'Game is no longer valid'}), 400
#     if this_station not in GAMES[this_game].network_stations.keys():
#         return jsonify({'error': 'Station is no longer valid'}), 400
#     if this_station not in GAMES[this_game].manual_stations_names:
#         return jsonify({'error': 'Station has been switched to autopilot'}), 400
#     if 'player_name' not in session.keys():
#         session['player_name'] = ''
#     if GAMES[this_game].network_stations[this_station].player_name == session['player_name']:
#         GAMES[this_game].network_stations[this_station].touch()
#     elif time.time() - GAMES[this_game].network_stations[this_station].last_communication_time > SecondsAway_to_Disconnect:
#         GAMES[this_game].network_stations[this_station].touch()
#         GAMES[this_game].network_stations[this_station].player_name = session['player_name']
#     else:
#         return jsonify({'error': 'Station already selected by another player'}), 400

#     static_info = {
#         'weeks': GAMES[this_game].weeks,
#         'current_week': GAMES[this_game].current_week,
#         'turn_time': GAMES[this_game].turn_time,
#         'number_of_players': len(GAMES[this_game].manual_stations_names),
#         'secondsaway_to_disconnect': SecondsAway_to_Disconnect,
#         'RefreshInterval_PlayScreen': RefreshInterval_PlayScreen,
#         'suppliers': [x.station_name for x in GAMES[this_game].network_stations[this_station].suppliers],
#         'customers': [x.station_name for x in GAMES[this_game].network_stations[this_station].customers],
#         'auto_ship': GAMES[this_game].network_stations[this_station].auto_decide_ship_qty,
#         'auto_order': GAMES[this_game].network_stations[this_station].auto_decide_order_qty,
#         'game_info': {
#             'holding_cost': GAMES[this_game].network_stations[this_station].holding_cost,
#             'backorder_cost': GAMES[this_game].network_stations[this_station].backorder_cost,
#             'transport_cost': GAMES[this_game].network_stations[this_station].transport_cost,
#             'transport_size': GAMES[this_game].network_stations[this_station].transport_size,
#             'delay_shipping': GAMES[this_game].network_stations[this_station].delay_shipping,
#             'delay_ordering': GAMES[this_game].network_stations[this_station].delay_ordering
#         }
#     }

#     if len(GAMES[this_game].network_stations[this_station].suppliers) == 0 and \
#        not GAMES[this_game].network_stations[this_station].auto_decide_order_qty:
#         static_info['suppliers'] = ['MyWorkshop']

#     return jsonify(static_info), 200


@app.route('/join_and_play', methods=['POST'])
@cross_origin()
def join_and_play():
    try:
        data = request.json
        selected_game = data.get('selected_game')
        player_name = data.get('player_name')
        password = data.get('password')
        selected_station = data.get('selected_station')

        if not all([selected_game, player_name, password, selected_station]):
            return jsonify({'error': 'Missing required parameters'}), 400

        if selected_game not in GAMES:
            return jsonify({'error': 'Selected game is no longer valid. Please select another game.'}), 400

        game = GAMES[selected_game]

        if game.play_password != password:
            return jsonify({'error': 'Incorrect password!'}), 400

        if selected_station not in game.network_stations:
            return jsonify({'error': 'Invalid station selected'}), 400

        station = game.network_stations[selected_station]

        if not station.player_name.startswith('player') and station.player_name != player_name:
            return jsonify({'error': 'Selected station is already taken'}), 400

        # Update the player name in the game
        station.player_name = player_name
        station.touch()

        # Update MongoDB
        result = db.games.update_one(
            {'_id': selected_game, 'stations.name': selected_station},
            {'$set': {
                'stations.$.player_name': player_name,
                'stations.$.last_communication_time': station.last_communication_time
            }}
        )
        if result.modified_count == 0:
            logger.warning(f"Failed to update player name in MongoDB for game {selected_game}, station {selected_station}")

        # Prepare play screen information
        static_info = {
            'weeks': game.weeks,
            'current_week': game.current_week,
            'turn_time': game.turn_time,
            'number_of_players': len(game.manual_stations_names),
            'secondsaway_to_disconnect': SecondsAway_to_Disconnect,
            'RefreshInterval_PlayScreen': RefreshInterval_PlayScreen,
            'suppliers': [x.station_name for x in station.suppliers],
            'customers': [x.station_name for x in station.customers],
            'auto_ship': station.auto_decide_ship_qty,
            'auto_order': station.auto_decide_order_qty,
            'game_info': {
                'holding_cost': station.holding_cost,
                'backorder_cost': station.backorder_cost,
                'transport_cost': station.transport_cost,
                'transport_size': station.transport_size,
                'delay_shipping': station.delay_shipping,
                'delay_ordering': station.delay_ordering
            }
        }

        if len(station.suppliers) == 0 and not station.auto_decide_order_qty:
            static_info['suppliers'] = ['MyWorkshop']

        return jsonify({
            'message': 'Successfully joined the game',
            'game': selected_game,
            'station': selected_station,
            'player_name': player_name,
            'game_info': static_info
        }), 200

    except pymongo.errors.PyMongoError as e:
        logger.exception(f"A MongoDB error occurred in join_and_play: {str(e)}")
        return jsonify({'error': f'A database error occurred: {str(e)}'}), 500
    except Exception as e:
        logger.exception(f"An error occurred in join_and_play: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    
@app.route('/monitor_games_screen', methods=['POST','GET'])
def monitor_games_screen():
    games_list = []
    actions_allowed = {}  # (T/F) for each
    if request.method == 'GET':  # when user reloads page
        pass
    if request.method == 'POST':
        for i in [x for x in range(MaxGamesMonitored)]:
            G = request.form.get('selected_game'+str(i))
            P = request.form.get('password'+str(i))
            if G not in games_list and G in GAMES.keys():
                games_list.append(G)
                if GAMES[G].admin_password == P:
                    actions_allowed[G] = True
                else:
                    actions_allowed[G] = False
    return render_template('multi-game-monitor.html', games_list=games_list, actions_allowed=actions_allowed,RefreshInterval=RefreshInterval_GameMonitor)


@app.route('/monitor_screen', methods=['GET'])
def monitor_screen():
    this_game = request.args.get('game_name')
    if this_game not in GAMES.keys():
        flash('Error: Game not found.')
        return redirect(url_for('admin_game'))
    static_info = {'game':this_game,
                   'weeks':GAMES[this_game].weeks,
                   'stations':GAMES[this_game].auto_stations_names + GAMES[this_game].manual_stations_names,
                   'demands':GAMES[this_game].demand_stations_names,
                   'station_players':{k:v.player_name for (k,v) in GAMES[this_game].network_stations.items()},
                   'RefreshInterval_GameMonitor':RefreshInterval_GameMonitor}
    return render_template('monitor_screen.html', static_info=static_info)


@app.route('/monitor_screen_radar', methods=['GET'])
def monitor_screen_radar():
    this_game = request.args.get('game_name')
    if this_game not in GAMES.keys():
        flash('Error: Game not found.')
        return redirect(url_for('admin_game'))
    static_info = {'game':this_game,
                   'stations':GAMES[this_game].auto_stations_names + GAMES[this_game].manual_stations_names,
                   'station_players':{k:v.player_name for (k,v) in GAMES[this_game].network_stations.items()},
                   'RefreshInterval_GameMonitor':RefreshInterval_GameMonitor}
    return render_template('monitor_screen_radar.html', static_info=static_info)


@app.route('/show_network', methods=['GET'])
def show_network():
    this_game = request.args.get('game_name')
    if this_game not in GAMES.keys():
        flash('Error: Game is no longer valid.')
        return render_template('network.html', static_info={'nodes':[], 'edges':[]})
    nodes = []
    node_id = {}
    i = 1
    for x in GAMES[this_game].demand_stations_names:
        nodes.append({'id': i, 'label': x, 'image':url_for('static', filename='clients.png')})
        node_id[x] = i
        i += 1
    for x in GAMES[this_game].manual_stations_names:
        nodes.append({'id': i, 'label': x + '\n(' + GAMES[this_game].network_stations[x].player_name + ')', 'image':url_for('static', filename='warehouse1.png')})
        node_id[x] = i
        i += 1
    for x in GAMES[this_game].auto_stations_names:
        nodes.append({'id': i, 'label': x + '\n(' + GAMES[this_game].network_stations[x].player_name + ')', 'image':url_for('static', filename='warehouse0.png')})
        node_id[x] = i
        i += 1
    edges = []
    config = GAMES[this_game].get_config()
    for x in config['connections']:
        edges.append({'from': node_id[x['supp']], 'to': node_id[x['cust']], 'arrows':'middle'})
    static_info = {'nodes':nodes, 'edges':edges}
    return render_template('network.html', static_info=static_info)


################################################################################
# game and station selection screens
################################################################################
def select_game_page(text,next_page,ask_name,ask_password):
    kill_expired_games()
    data = {'text':text,
            'next_page':next_page,
            'ask_name':ask_name,
            'ask_password':ask_password,
            'games_list':[x.team_name for x in GAMES.values()]}
    return render_template('select_game.html', data=data)


def select_station_page(text,game_name,next_page):
    if game_name not in GAMES.keys():
        flash('Game is no longer valid, please try again.')
        return redirect(url_for('cmd',cmd='join_game'))
    data = {'text':text,
            'next_page':next_page,
            'stations_list':[x for x in GAMES[game_name].manual_stations_names
                             if time.time() - GAMES[game_name].network_stations[x].last_communication_time > SecondsAway_to_Disconnect]}
    return render_template('select_station.html', data=data)

def infinity_to_string(obj):
    if isinstance(obj, dict):
        return {k: infinity_to_string(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [infinity_to_string(v) for v in obj]
    elif isinstance(obj, float) and math.isinf(obj):
        return "Infinity"
    return obj

################################################################################
# game data management
################################################################################

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, float) and math.isinf(obj):
            return "Infinity" if obj > 0 else "-Infinity"
        return super().default(obj)

@app.route('/edit_game_setup', methods=['POST', 'GET'])
@cross_origin()
def edit_game_setup():
    action = request.args.get('action', '')
    this_game = request.args.get('game')
    
    logger.info(f"edit_game_setup called with method {request.method}, action: {action}, game: {this_game}")

    if request.method == 'GET':
        if not this_game or this_game not in GAMES.keys():
            logger.warning(f"Invalid game requested: {this_game}")
            return jsonify({'error': 'Game is no longer valid. Please select another game.'}), 400
        
        try:
            game_config = GAMES[this_game].get_config()
            # Use the infinity_to_string function to preprocess the game configuration
            game_config_processed = infinity_to_string(game_config)
            return jsonify({'setup': game_config_processed, 'action': action}), 200
        except Exception as e:
            logger.exception(f"Error getting game config for {this_game}: {str(e)}")
            return jsonify({'error': f"Error retrieving game configuration: {str(e)}"}), 500

    elif request.method == 'POST':
        if not this_game:
            return jsonify({'error': 'No game selected.'}), 400

        if action == 'edit':
            if this_game not in GAMES.keys():
                return jsonify({'error': 'Game is no longer valid. Please select another game.'}), 400
            # Handle edit action here
            return jsonify({'message': 'Game setup edited successfully', 'selected_game': this_game}), 200

        elif action == 'create':
            kill_expired_games()
            filename = f'./game_templates/{this_game}.json'
            if not os.path.isfile(filename) or not allowed_file(filename):
                logger.warning(f"Template not found or invalid: {filename}")
                return jsonify({'error': 'Template not found or is not a JSON file.'}), 400
            
            try:
                with open(filename) as game_file:
                    game_data = json.load(game_file)
                i = 1
                new_game_name = f"{game_data['team_name']}{i}"
                while new_game_name in GAMES.keys():
                    i += 1
                    new_game_name = f"{game_data['team_name']}{i}"
                game_data['team_name'] = new_game_name
                GAMES[new_game_name] = game.Game(game_data)
                session['selected_game'] = new_game_name
                logger.info(f"New game created: {new_game_name}")
                return jsonify({'message': 'Game created successfully', 'selected_game': new_game_name}), 201
            except Exception as e:
                logger.exception(f"Error creating game: {str(e)}")
                if 'new_game_name' in locals() and new_game_name in GAMES.keys():
                    del GAMES[new_game_name]
                return jsonify({'error': f"Could not create game: {str(e)}"}), 500

        else:
            logger.warning(f"Invalid action requested: {action}")
            return jsonify({'error': 'No valid action provided; must include edit or create actions.'}), 400

    logger.error(f"Unexpected request method: {request.method}")
    return jsonify({'error': 'Invalid request method'}), 405

@app.route('/import_game_file', methods=['POST'])
def import_game_file():
    if 'importfilename' not in request.files:
        flash('No file part in posted request.')
        return redirect(request.url)
    file = request.files['importfilename']
    if file.filename == '':
        flash('No selected file.')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        game_data = json.loads(file.read().decode('ascii'))
    else:
        flash('Can\'t load file.')
        return redirect(request.url)
    kill_expired_games()
    if game_data['team_name'] in GAMES.keys():
        flash(game_data['team_name'] + ': Game already exists.')
    else:
        try:
            GAMES[game_data['team_name']] = game.Game(game_data)
            flash(game_data['team_name'] + ': Game successfully added.')
        except Exception as e:
            if game_data.get('team_name','') in GAMES.keys():
                del GAMES[game_data['team_name']]
            show_error("Could not create game",e)
    return redirect(url_for('cmd',cmd='import_game'))


@app.route('/export_game_file', methods=['GET'])
def export_game_file():
    game_name = request.args.get('game_name','')
    if game_name not in GAMES.keys():
        flash('Error: Game is no longer valid.')
        return redirect(url_for('admin_game'))
    game_file = json.dumps(GAMES[game_name].get_config(),cls=MyEncoder,indent=2,sort_keys=True)
    return Response(
        game_file,
        mimetype="application/json",
        headers={"Content-disposition":"attachment; filename=game_export.json"})


@app.route('/show_game_setup', methods=['GET'])
def show_game_setup():
    game_name = request.args.get('game_name','')
    if game_name not in GAMES.keys():
        flash('Error: Game is no longer valid.')
        return redirect(url_for('admin_game'))
    else:
        try:
            game_setup = GAMES[game_name].get_config()
            return render_template('show_game_setup.html', setup=game_setup, sort_keys=True)
        except Exception as e:
            show_error("Couldn't get game setup data",e)
            return redirect(url_for('admin_game'))


@app.route('/copy_game', methods=['POST'])
def copy_game():
    org = request.args.get('org','')
    dst = request.form.get('dst','')
    if dst == '':
        flash('Error: Please provide a new game name.')
        return redirect(url_for('admin_game'))
    if org not in GAMES.keys():
        flash('Error: Origin game is no longer valid.')
        return redirect(url_for('admin_game'))
    if dst in GAMES.keys():
        flash('Error: destination game already exists.')
        return redirect(url_for('admin_game'))
    try:
        game_config = GAMES[org].get_config()
        game_config['team_name'] = dst
        GAMES[dst] = game.Game(game_config)
        flash(dst + ': Game successfully copied.')
    except Exception as e:
        if dst in GAMES.keys():
            del GAMES[dst]
        show_error("Could not create game", e)
    return redirect(url_for('admin_game'))


@app.route('/rename_game', methods=['POST'])
def rename_game():
    org = request.args.get('org','')
    dst = request.form.get('new_name','')
    if dst == '':
        flash('Error: Please provide a new game name.')
        return redirect(url_for('admin_game'))
    if org not in GAMES.keys():
        flash('Error: Origin game is no longer valid.')
        return redirect(url_for('admin_game'))
    if dst in GAMES.keys():
        flash('Error: destination game already exists.')
        return redirect(url_for('admin_game'))
    try:
        GAMES[dst] = GAMES.pop(org)
        GAMES[dst].team_name = dst
        session['selected_game'] = dst
        flash(dst + ': Game successfully renamed.')
    except Exception as e:
        show_error("Could not rename game",e)
    return redirect(url_for('admin_game'))


@app.route('/delete_game', methods=['GET'])
def delete_game():
    game_name = request.args.get('game_name','')
    next_page = request.args.get('next','advanced_menu')
    if game_name in GAMES.keys():
        del GAMES[game_name]
        flash(game_name + ': Game deleted.')
    else:
        flash('Error: Game not found.')
    return redirect(url_for('cmd', cmd=next_page))


@app.route('/reset_game', methods=['GET','POST'])
def reset_game():
    game_name = request.args.get('game_name','')
    if request.method == 'POST':  # when called from $.post()
        if game_name in GAMES.keys():
            try:
                GAMES[game_name].reset()
            except Exception as e:
                return jsonify({'Result': '{0}:{1!r}'.format(type(e).__name__, e.args)})
            return jsonify({'Result': ''})  # game_name + ' was successfully reset.'}) # # reset effects are obvious on the screen, no need for a message
        else:
            return jsonify({'Result': game_name + ' is no longer valid.'})
    if request.method == 'GET':  # when redirected ehre or user reloads page
        if game_name not in GAMES.keys():
            flash('Error: Game is no longer valid.')
            return redirect(url_for('admin_game'))
        else:
            try:
                GAMES[game_name].reset()
                flash(game_name + ': Game successfully reset.')
            except Exception as e:
                show_error("Could not reset game",e)
        return redirect(url_for('admin_game'))


@app.route('/save_games_state', methods=['GET'])
def save_server_state():
    if save_games_state():
        flash('Games states saved.')
    else:
        flash('Error: Could not save games state.')
    return redirect(url_for('admin_server_menu'))


@app.route('/load_games_state', methods=['GET'])
def load_server_state():
    if load_games_state():
        flash('Games states loaded.')
    else:
        flash('Error: Could not load games state.')
    return redirect(url_for('admin_server_menu'))


################################################################################
# Debug screens
################################################################################
#
# @app.route('/debug_wsg', methods=['GET'])
# def debug_wsg():
#     # return GAMES['WSG_test'].Debug_Report_WSG_Inv_PO_Report()
#     if 'WSG' in GAMES.keys():
#         return '<!DOCTYPE html><html><head><title>WSG Debug Screen</title></head>' + \
#                 '<body><pre>' + GAMES['WSG'].Debug_Report_WSG_Inv_PO_Report() + \
#                 '</pre></body></html>'
#     else:
#         flash('No such game: WSG')
#         return redirect(url_for('index'))


################################################################################
# AJAX
################################################################################
@app.route('/change_game_settings', methods=['POST'])
@cross_origin()
def change_game_settings():
    try:
        this_game = request.args.get('game')
        logger.info(f"Received request to change settings for game: {this_game}")

        if this_game not in GAMES:
            return jsonify({'result': False, 'msg': f"Game {this_game} not found."}), 404

        # Log request details for debugging
        logger.debug(f"Request headers: {dict(request.headers)}")
        logger.debug(f"Request method: {request.method}")
        logger.debug(f"Request content type: {request.content_type}")
        logger.debug(f"Raw request data: {request.get_data(as_text=True)}")

        try:
            game_settings = request.get_json(force=True)
            if game_settings is None:
                raise ValueError("Empty JSON data")
        except Exception as json_error:
            logger.error(f"JSON parsing error: {str(json_error)}")
            # Attempt to read raw data
            raw_data = request.get_data(as_text=True)
            logger.debug(f"Raw request data: {raw_data}")
            return jsonify({
                'result': False,
                'msg': f"Failed to parse JSON data: {str(json_error)}. Raw data: {raw_data[:200]}..."
            }), 400

        logger.debug(f"Parsed game settings: {game_settings}")

        # Convert text lists to numeric lists
        logger.debug("Processing demand settings")
        for x in game_settings.get('demands', []):
            x['demand'] = txt2list(x['demand'])

        logger.debug("Processing station settings")
        for x in game_settings.get('stations', []):
            x['order_max'] = txt2list(x['order_max'])
            x['order_min'] = txt2list(x['order_min'])
            x['ship_max'] = txt2list(x['ship_max'])
            x['ship_min'] = txt2list(x['ship_min'])

        logger.debug("Processing script settings")
        for x in game_settings.get('script', []):
            x['week'] = int(x['week'])

        # Ensure that the game name is not changed
        game_settings['team_name'] = this_game
        game_settings['_id'] = this_game  # Use the game name as MongoDB _id

        # Create a new game object with the updated settings
        logger.info("Creating new game object with updated settings")
        try:
            G = game.Game(game_settings)
        except Exception as game_error:
            logger.error(f"Error creating game object: {str(game_error)}")
            return jsonify({'result': False, 'msg': f"Error in game settings: {str(game_error)}"}), 400

        # Update the game in the GAMES dictionary
        logger.info(f"Updating game: {this_game}")
        GAMES[this_game] = G

        # Update the game in MongoDB
        try:
            result = db.games.replace_one({'_id': this_game}, game_settings, upsert=True)
            if result.modified_count > 0 or result.upserted_id:
                logger.info(f"Game {this_game} updated in MongoDB")
            else:
                logger.warning(f"No changes detected for game {this_game} in MongoDB")
        except pymongo.errors.PyMongoError as mongo_error:
            logger.error(f"MongoDB error: {str(mongo_error)}")
            return jsonify({'result': False, 'msg': f"Database error: {str(mongo_error)}"}), 500

        logger.info(f"Game settings updated successfully for: {this_game}")
        return jsonify({'result': True, 'msg': 'Game setup saved successfully.'})

    except Exception as e:
        logger.error(f"Error occurred while saving game setup for {this_game}: {str(e)}", exc_info=True)
        return jsonify({
            'result': False,
            'msg': f"Couldn't save game setup. Error: {str(e)}"
        }), 500  
    
@app.route('/submit', methods=['POST'])
@cross_origin()
def submit():
    try:
        data = request.json
        this_game = data['DATA']['this_game']
        this_station = data['DATA']['this_station']

        if not this_game or not this_station or this_game not in GAMES or this_station not in GAMES[this_game].network_stations:
            return jsonify({'error': 'Invalid game or station'}), 400

        # Get JSON data from the request
        this_week = data['DATA']['week']
        suppliers = data['DATA']['suppliers']
        customers = data['DATA']['customers']

        # Convert text input into numbers
        suppliers = {k: int(v) for k, v in suppliers.items()}
        customers = {k: int(v) for k, v in customers.items()}

        # Update game state
        GAMES[this_game].network_stations[this_station].touch()
        GAMES[this_game].SetPlayerTurnData(this_station, this_week, suppliers, customers)

        return jsonify({'message': 'Turn data submitted successfully'}), 200
    except Exception as e:
        log_msg("Issue in player submission. Game: {0}. Exception: {1} - {2!r}".format(this_game, type(e).__name__, e.args))
        return jsonify({'error': 'An error occurred while processing the request'}), 500


@app.route('/get_station_status', methods=['GET'])
def get_station_status():
    this_game = request.args.get('game','')
    this_station = request.args.get('station','')

    if this_game not in GAMES.keys():
        log_msg("'{:}' is no longer a valid game. AJAX get_station_status request ignored.".format(this_game))
        return jsonify({})
    if this_station not in GAMES[this_game].network_stations.keys():
        log_msg("'{:}.{:}' is no longer a valid station. AJAX get_station_status request ignored.".format(this_game, this_station))
        return jsonify({})

    GAMES[this_game].network_stations[this_station].touch()
    w = GAMES[this_game].current_week
    if w >= GAMES[this_game].weeks:
        w = GAMES[this_game].weeks - 1
    current_time = time.time()
    GAMES[this_game].connected_stations = len([1 for x in GAMES[this_game].network_stations.values() if (current_time - x.last_communication_time) < SecondsAway_to_Disconnect])

    station = GAMES[this_game].network_stations[this_station]
    incoming_shipment = show_week(station.outbound, w + 1) if w + 1 < GAMES[this_game].weeks else {}
    
    # New approach to determine shipments
    if station.auto_decide_ship_qty:
        shipments = station.decide_shipment(w)
    else:
        shipments = station.player_shipment[-1] if station.player_shipment else {}

    return jsonify({
        'current_week': w,
        'server_time': current_time,
        'script': next((i for i in GAMES[this_game].script if i['week'] == w + 1), {'msg': ''})['msg'],
        'turn_start_time': GAMES[this_game].turn_start_time,
        'game_done': GAMES[this_game].game_done,
        'players_completed_turn': GAMES[this_game].players_completed_turn,
        'connected_stations': GAMES[this_game].connected_stations,
        'connection_state': station.last_communication_time,
        'cost_inventory': game.currency(sum(station.kpi_weeklycost_inventory)),
        'cost_backorder': game.currency(sum(station.kpi_weeklycost_backorder)),
        'cost_transport': game.currency(sum(station.kpi_weeklycost_transport)),
        'total_cost': game.currency(sum(station.kpi_total_cost)),
        'fulfilment_rate': game.percent(list_avg(station.kpi_fulfilment_rate[:w])),
        'truck_utilization': game.percent(list_avg(station.kpi_truck_utilization[:w])),
        'inventory': station.inventory[w],
        'backorder': show_week(station.backorder, w),
        'incomming_shipment': incoming_shipment,
        'incomming_order': show_week(station.received_po, w),
        'incomming_delivery': show_week(station.inbound, w),
        'order_min': str(station.order_min[w]),
        'order_max': str(station.order_max[w]),
        'ship_min': str(station.ship_min[w]),
        'ship_max': str(station.ship_max[w]),
        'ordering_limits': toHTMLtbl([('wk', 'min', 'max')] + station.ordering_limits[w + 1:w + 11]),
        'shipping_limits': toHTMLtbl([('wk', 'min', 'max')] + station.shipping_limits[w + 1:w + 11]),
        'shipments': shipments  # Using the new shipments determination
    })


@cross_origin()
@app.route('/get_game_status', methods=['GET'])
def get_game_status():
    this_game = request.args.get('game','')
    if this_game not in GAMES.keys():
        log_msg("'{:}' is no longer a valid game. AJAX get_game_status request ignored.".format(this_game))
        return jsonify({})
    current_week = GAMES[this_game].current_week
    current_time = time.time()

    if current_week == 0:
        return jsonify({
            'week': current_week,
            'waitingfor': [GAMES[this_game].network_stations[x].station_name for x in GAMES[this_game].manual_stations_names
                           if GAMES[this_game].network_stations[x].week_turn_completed < current_week],
            'disconnected': [GAMES[this_game].network_stations[x].station_name for x in GAMES[this_game].manual_stations_names
                             if (current_time - GAMES[this_game].network_stations[x].last_communication_time) >= SecondsAway_to_Disconnect]
        })

    stations_list = GAMES[this_game].manual_stations_names + GAMES[this_game].auto_stations_names
    demands_list = GAMES[this_game].demand_stations_names
    cost_inventory_sum = {k: [0] for k in stations_list}
    cost_backorder_sum = {k: [0] for k in stations_list}
    cost_transport_sum = {k: [0] for k in stations_list}
    cost_total_sum = {k: [0] for k in stations_list}
    fulfilment_avg = {k: [0] for k in stations_list}
    green_score_avg = {k: [0] for k in stations_list}
    cost = {k: [0]*current_week for k in stations_list}
    fulfilment = {k: [0]*current_week for k in stations_list}
    green_score = {k: [0]*current_week for k in stations_list}
    shipments = {k: [0]*current_week for k in stations_list}
    extra_shipments = {k: [0]*current_week for k in stations_list}
    backorder = {k: [] for k in stations_list}
    inventory = {k: [] for k in stations_list}
    orders = {k: [] for k in (stations_list + demands_list)}
    deliveries = {k: [] for k in stations_list}

    player_info = {}
    player_performance = {}

    for x in (stations_list + demands_list):
        S = GAMES[this_game].network_stations[x]
        player_name = S.player_name if hasattr(S, 'player_name') else 'Demand'
        
        # Check if it's a Station or Demand object
        if hasattr(S, 'auto_decide_order_qty'):
            is_cpu = S.auto_decide_order_qty
        else:
            is_cpu = True  # Demand objects are always considered "CPU"

        player_info[x] = {
            'name': player_name,
            'is_cpu': is_cpu,
            'station': x
        }

        if x in GAMES[this_game].demand_stations_names:
            orders[x] = S.demand
        else:
            cost_inventory_sum[x] = sum(S.kpi_weeklycost_inventory[:current_week])
            cost_backorder_sum[x] = sum(S.kpi_weeklycost_backorder[:current_week])
            cost_transport_sum[x] = sum(S.kpi_weeklycost_transport[:current_week])
            cost_total_sum[x] = sum(S.kpi_total_cost[:current_week])
            fulfilment_avg[x] = sum(S.kpi_fulfilment_rate[:current_week])/(current_week)
            green_score_avg[x] = sum(S.kpi_truck_utilization[:current_week])/(current_week)
            cost[x] = S.kpi_total_cost
            fulfilment[x] = S.kpi_fulfilment_rate
            green_score[x] = S.kpi_truck_utilization
            shipments[x] = S.kpi_shipment_trucks
            for w in range(current_week):
                extra_shipments[x][w] = floor(S.kpi_shipment_trucks[w]*(1-S.kpi_fulfilment_rate[w]))
            backorder[x] = game.combine_weekly(S.backorder)
            inventory[x] = S.inventory
            orders[x] = game.combine_weekly(S.sent_po)
            deliveries[x] = game.combine_weekly(S.outbound)

            player_performance[x] = {
                'cost_inventory': cost_inventory_sum[x],
                'cost_backorder': cost_backorder_sum[x],
                'cost_transport': cost_transport_sum[x],
                'cost_total': cost_total_sum[x],
                'fulfilment_rate': fulfilment_avg[x],
                'green_score': green_score_avg[x],
                'weekly_cost': cost[x],
                'weekly_fulfilment': fulfilment[x],
                'weekly_green_score': green_score[x],
                'weekly_shipments': shipments[x],
                'weekly_extra_shipments': extra_shipments[x],
                'backorder': backorder[x],
                'inventory': inventory[x],
                'orders': orders[x],
                'deliveries': deliveries[x]
            }

    return jsonify({
        'week': current_week,
        'waitingfor': [GAMES[this_game].network_stations[x].station_name for x in GAMES[this_game].manual_stations_names
                       if GAMES[this_game].network_stations[x].week_turn_completed < current_week],
        'disconnected': [GAMES[this_game].network_stations[x].station_name for x in GAMES[this_game].manual_stations_names
                         if (current_time - GAMES[this_game].network_stations[x].last_communication_time) >= SecondsAway_to_Disconnect],
        'players': player_info,
        'player_performance': player_performance,
        'data': {
            'cost_inventory_sum': cost_inventory_sum,
            'cost_backorder_sum': cost_backorder_sum,
            'cost_transport_sum': cost_transport_sum,
            'cost_total_sum': cost_total_sum,
            'fulfilment_avg': fulfilment_avg,
            'green_score_avg': green_score_avg,
            'cost': cost,
            'fulfilment': fulfilment,
            'green_score': green_score,
            'shipments': shipments,
            'extra-shipments': extra_shipments,
            'backorder': backorder,
            'inventory': inventory,
            'orders': orders,
            'deliveries': deliveries
        }
    })


@app.route('/get_game_status_radar', methods=['GET'])
def get_game_status_radar():
    data = {}
    best = {'inventory':float("inf"),'backorder':float("inf"),'transport':float("inf"),'fulfilment':0,'green':0}
    this_game = request.args.get('game','')
    if this_game not in GAMES.keys():
        log_msg("'{:}' is no longer a valid game. AJAX get_game_status_radar request ignored.".format(this_game))
        return jsonify({'nodata': True})
    current_week = GAMES[this_game].current_week
    stations_list = GAMES[this_game].manual_stations_names + GAMES[this_game].auto_stations_names
    if current_week == 0:
        return jsonify({'nodata':True})

    for x in stations_list:
        S = GAMES[this_game].network_stations[x]
        data[x] = {'inventory':sum(S.kpi_weeklycost_inventory[:current_week]),
                   'backorder':sum(S.kpi_weeklycost_backorder[:current_week]),
                   'transport':sum(S.kpi_weeklycost_transport[:current_week]),
                   'fulfilment':sum(S.kpi_fulfilment_rate[:current_week])/(current_week),
                   'green':sum(S.kpi_truck_utilization[:current_week])/(current_week)}
        for k,v in data[x].items():
            if k in ['inventory','backorder','transport']:
                if v < best[k]:
                    best[k] = v
            if k in ['fulfilment','green']:
                if v > best[k]:
                    best[k] = v
    for x in stations_list:
        for k,v in data[x].items():
            if k in ['fulfilment','green']:
                if best[k] > 0:
                    data[x][k] = data[x][k]/best[k]
            if k in ['inventory','backorder','transport']:
                if data[x][k] > 0:
                    data[x][k] = best[k]/data[x][k]
    return jsonify(data)


@app.route('/get_games_status', methods=['GET'])
def get_games_status():
    games_list = request.args.get('games_list','').split(',')
    current_time = time.time()
    data = {}
    for g in games_list:
        if g not in GAMES.keys():
            log_msg("'{:}' is no longer a valid game. Game skipped in AJAX get_games_status request.".format(g))
        else:
            current_week = GAMES[g].current_week
            if current_week >= GAMES[g].weeks:
                current_week = GAMES[g].weeks - 1

            # list disconnected/waiting stations, making sure missing stations are replaced with an empty entry for sake of constant table size creation
            waiting_list = [(GAMES[g].network_stations[x].station_name,) for x in GAMES[g].manual_stations_names if GAMES[g].network_stations[x].week_turn_completed < current_week]
            waiting_list.extend((('---',),)*(len(GAMES[g].manual_stations_names)-len(waiting_list)))
            disconnected_list = [(GAMES[g].network_stations[x].station_name,) for x in GAMES[g].manual_stations_names if (current_time - GAMES[g].network_stations[x].last_communication_time) >= SecondsAway_to_Disconnect]
            disconnected_list.extend((('---',),)*(len(GAMES[g].manual_stations_names)-len(disconnected_list)))

            if current_week == 0:
                data[g] = {'week':current_week+1,
                           'waiting':toHTMLtbl(waiting_list).replace(',',''),
                           'disconnected':toHTMLtbl(disconnected_list).replace(',',''),
                           'inventory':0,'backorder':0,'trucks':0,'cost':0,'fulfilment':0,'greenscore':0,'totalcost':[0],'avgfulfilment':[0],'avggreenscore':[0]}
            else:
                if GAMES[g].game_done:
                    cost = GAMES[g].kpi_cost['total']
                    fulfilment = GAMES[g].kpi_customer_satisfaction
                    greenscore = GAMES[g].kpi_green_score
                else:
                    cost = GAMES[g].kpi_cost['total'][:current_week]
                    fulfilment = GAMES[g].kpi_customer_satisfaction[:current_week]
                    greenscore = GAMES[g].kpi_green_score[:current_week]
                data[g] = {'week':current_week+1,
                           'waiting':toHTMLtbl(waiting_list).replace(',',''),
                           'disconnected':toHTMLtbl(disconnected_list).replace(',',''),
                           'inventory':sum([x.inventory[current_week] for x in GAMES[g].network_stations.values() if type(x) is not game.stations.Demand]),
                           'backorder':sum([week_sum(x.backorder,current_week) for x in GAMES[g].network_stations.values() if type(x) is not game.stations.Demand]),
                           'trucks':sum(GAMES[g].kpi_trucks),
                           'cost':sum(cost),
                           'fulfilment':list_avg(fulfilment),
                           'greenscore':list_avg(greenscore),
                           'totalcost':[numpy.asscalar(x) for x in numpy.cumsum(cost)],
                           'avgfulfilment':list_cum_avg(fulfilment),
                           'avggreenscore':list_cum_avg(greenscore)}
    return jsonify(data)


@app.route('/get_debug_data', methods=['GET'])
def get_debug_data():
    this_game = request.args.get('game','')
    if this_game not in GAMES.keys():
        log_msg("'{:}' is no longer a valid game. AJAX get_debug_data request ignored.".format(this_game))
        return jsonify({})
    w = GAMES[this_game].current_week
    current_time = time.time()
    GAMES[this_game].connected_stations = len([1 for x in GAMES[this_game].network_stations.values() if (current_time - x.last_communication_time) < SecondsAway_to_Disconnect])
    return jsonify({
        'week':w,
        'players_completed_turn':GAMES[this_game].players_completed_turn,
        'connected_stations':GAMES[this_game].connected_stations,
        'report':GAMES[this_game].Debug_Report()})

# Sessions
@app.route('/create_session', methods=['POST'])
@cross_origin()
def create_session():
    logger.info("Entering create_session function")
    try:
        data = request.json
        logger.debug(f"Received data: {data}")

        session_name = data.get('session_name')
        num_teams = data.get('num_teams')
        user_id = data.get('user_id')

        logger.info(f"Extracted session_name: {session_name}, num_teams: {num_teams}, user_id: {user_id}")

        if not session_name or not num_teams or not user_id:
            logger.error("Missing required parameters")
            return jsonify({'error': 'Missing session_name, num_teams, or user_id'}), 400

        try:
            num_teams = int(num_teams)
            if num_teams < 1:
                raise ValueError
        except ValueError:
            logger.error(f"Invalid num_teams value: {num_teams}")
            return jsonify({'error': 'num_teams must be a positive integer'}), 400

        # Load the Root Beer Game template
        template_filename = './game_templates/Root Beer Game.json'
        logger.info(f"Loading game template from {template_filename}")
        if not os.path.exists(template_filename):
            logger.error(f"Template file not found: {template_filename}")
            return jsonify({'error': 'Game template not found'}), 500

        with open(template_filename, 'r') as template_file:
            template_data = json.load(template_file)

        logger.info("Creating new Session object")
        new_session = Session(session_name, user_id, num_teams)
        session_id = new_session.id
        logger.info(f"Created Session with ID: {session_id}")

        logger.info("Creating or updating games for each team")
        for i in range(num_teams):
            logger.debug(f"Processing game {i+1} of {num_teams}")
            game_name = f"{session_name}_Team_{i+1}"

            # Create a deep copy of the template data
            game_data = json.loads(json.dumps(template_data))

            # Update the game data with session-specific information
            game_data.update({
                '_id': game_name,
                'team_name': game_name,
                'session_id': session_id,
                'team_number': i + 1,
                'play_password': 'play',
                'admin_password': 'admin',
                'status': 'created',
                'players': []
            })

            logger.debug(f"Adding game {game_name} to session's game list")
            new_session.games.append(game_name)

            logger.debug(f"Upserting game {game_name} in MongoDB")
            db.games.update_one({'_id': game_name}, {'$set': game_data}, upsert=True)

            logger.debug(f"Creating or updating game instance in GAMES dictionary")
            GAMES[game_name] = game.Game(game_data)

        logger.info("Saving session data to MongoDB")
        db.sessions.insert_one(new_session.to_dict())

        logger.info("Adding session to SESSIONS dictionary")
        SESSIONS[session_id] = new_session

        logger.info("Session creation completed successfully")
        session_link = f"{request.host_url}join_session/{session_id}"
        return jsonify({
            'message': 'Session created successfully',
            'session_id': session_id,
            'session_link': session_link,
            'num_games_created': num_teams
        }), 201

    except pymongo.errors.PyMongoError as e:
        logger.exception(f"A MongoDB error occurred: {str(e)}")
        return jsonify({'error': f'A database error occurred: {str(e)}'}), 500
    except Exception as e:
        logger.exception(f"An error occurred: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
# Helper function to create a game
def create_game(game_config):
    try:
        game_id = str(uuid.uuid4())
        
        # Default game configuration
        default_config = {
            'team_name': game_config['team_name'],
            'play_password': game_config['play_password'],
            'admin_password': game_config['admin_password'],
            'weeks': 52,  # Default to 52 weeks, adjust as needed
            'turn_time': 300,  # Default to 300 seconds, adjust as needed
            'expiry': 7,  # Default to 7 days, adjust as needed
            'demands': [
                {
                    'station_name': 'Customer',
                    'demand': [10] * 52,  # Default demand of 10 for each week
                    'auto_decide_po_qty': True
                }
            ],
            'stations': [
                {
                    'station_name': 'Manufacturer',
                    'inventory': 0,
                    'order_max': [100] * 52,
                    'order_min': [0] * 52,
                    'ship_max': [100] * 52,
                    'ship_min': [0] * 52,
                    'holding_cost': 1,
                    'backorder_cost': 2,
                    'transport_cost': 10,
                    'transport_size': 5,
                    'delay_shipping': 1,
                    'delay_ordering': 2,
                    'auto_decide_order_qty': False,
                    'auto_decide_ship_qty': False
                }
            ],
            'connections': [
                {
                    'supp': 'Manufacturer',
                    'cust': 'Customer'
                }
            ],
            'script': []  # Add game events/scripts if needed
        }
        
        # Merge the provided config with the default config
        full_config = {**default_config, **game_config}
        
        # Add the game_id as _id for MongoDB
        full_config['_id'] = game_id
        
        # Create game instance
        GAMES[game_id] = game.Game(full_config)
        
        # Save game to MongoDB
        result = db.games.insert_one(full_config)
        
        if result.inserted_id:
            logger.info(f"Game {game_id} created and saved to MongoDB")
            return game_id
        else:
            logger.error(f"Failed to save game {game_id} to MongoDB")
            del GAMES[game_id]  # Remove from memory if database insert failed
            return None

    except pymongo.errors.PyMongoError as e:
        logger.error(f"MongoDB error occurred while creating game: {str(e)}")
        if game_id in GAMES:
            del GAMES[game_id]  # Remove from memory if database insert failed
        return None

    except Exception as e:
        logger.error(f"Error occurred while creating game: {str(e)}")
        if game_id in GAMES:
            del GAMES[game_id]  # Remove from memory if an error occurred
        return None
    
# Get all sessions
@app.route('/get_sessions', methods=['GET'])
@cross_origin()
def get_sessions():
    sessions = [session.to_dict() for session in SESSIONS.values()]
    return jsonify(sessions), 200

# Join a session
@app.route('/join_session/<session_id>', methods=['POST'])
@cross_origin()
def join_session(session_id):
    data = request.json
    player_uid = data.get('player_uid')
    player_name = data.get('player_name')

    if not session_id or not player_uid or not player_name:
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        if session_id not in SESSIONS:
            return jsonify({'error': 'Session not found'}), 404

        session = SESSIONS[session_id]
        
        # Add the player to the session
        session.players.add(player_uid)

        # Update session in MongoDB
        result = db.sessions.update_one(
            {'_id': session_id},
            {'$addToSet': {'players': {'uid': player_uid, 'name': player_name, 'role': None, 'team': None}}},
            upsert=True
        )

        if result.modified_count == 0 and result.upserted_id is None:
            logger.error(f"Failed to update session {session_id} with player {player_uid}")
            return jsonify({'error': 'Failed to update session'}), 500

        return jsonify({'message': 'Joined session successfully', 'waiting_room_url': f"/wait_for_game_start/{session_id}/{player_uid}"}), 200

    except Exception as e:
        logger.exception(f"An error occurred in join_session: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

# Wait for game to start
@app.route('/wait_for_game_start/<session_id>/<player_uid>', methods=['GET'])
@cross_origin()
def wait_for_game_start(session_id, player_uid):
    try:
        if session_id not in SESSIONS:
            logger.error(f"Session {session_id} not found in SESSIONS dictionary")
            return jsonify({'error': 'Session not found'}), 404

        session = SESSIONS[session_id]
        
        session_data = db.sessions.find_one({'_id': session_id})
        
        if session_data is None:
            logger.error(f"Session {session_id} not found in database")
            return jsonify({'error': 'Session data not found in database'}), 404

        players = session_data.get('players', [])
        player = next((p for p in players if p['uid'] == player_uid), None)

        if player is None:
            logger.error(f"Player {player_uid} not found in session {session_id}")
            return jsonify({'error': 'Player not found in session'}), 404

        if player.get('role') and player.get('team'):
            game = GAMES.get(player['team'])
            if game:
                team_name = game.team_name
            else:
                team_name = "Unknown Team"
            return jsonify({'status': 'game_started', 'role': player['role'], 'team': player['team'], 'team_name': team_name}), 200
        else:
            return jsonify({'status': 'waiting'}), 200

    except Exception as e:
        logger.exception(f"An error occurred in wait_for_game_start: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    
# Get waiting students
@app.route('/get_waiting_players/<session_id>', methods=['GET'])
@cross_origin()
def get_waiting_players(session_id):
    try:
        if session_id not in SESSIONS:
            logger.warning(f"Session {session_id} not found in SESSIONS dictionary")
            return jsonify({'error': 'Session not found'}), 404

        session_data = db.sessions.find_one({'_id': session_id})
        if not session_data:
            logger.warning(f"Session {session_id} not found in database")
            return jsonify({'error': 'Session data not found in database'}), 404

        players = session_data.get('players', [])
        waiting_players = [p for p in players if p.get('role') is None or p.get('team') is None]

        return jsonify({'waiting_players': waiting_players}), 200

    except pymongo.errors.PyMongoError as e:
        logger.exception(f"A MongoDB error occurred: {str(e)}")
        return jsonify({'error': f'A database error occurred: {str(e)}'}), 500
    except Exception as e:
        logger.exception(f"An unexpected error occurred: {str(e)}")
        return jsonify({'error': f'An unexpected error occurred: {str(e)}'}), 500
# Assign Roles
@app.route('/assign_roles/<session_id>', methods=['POST'])
@cross_origin()
def assign_roles(session_id):
    try:
        if session_id not in SESSIONS:
            return jsonify({'error': 'Session not found'}), 404

        session = SESSIONS[session_id]
        
        # Check if roles have already been assigned
        session_data = db.sessions.find_one({'_id': session_id})
        if session_data.get('roles_assigned', False):
            return jsonify({'error': 'Roles have already been assigned for this session'}), 400

        data = request.json
        assignments = data.get('assignments', {})

        # Fetch current session data
        current_players = {p['uid']: p for p in session_data.get('players', [])}

        # Prepare bulk write operations
        bulk_operations = []

        for player_uid, assignment in assignments.items():
            new_role = assignment.get('role')
            new_team = assignment.get('team')

            update = {
                'players.$.role': new_role,
                'players.$.team': new_team
            }

            bulk_operations.append(
                pymongo.UpdateOne(
                    {'_id': session_id, 'players.uid': player_uid},
                    {'$set': update}
                )
            )

        # Add operation to set roles_assigned to True
        bulk_operations.append(
            pymongo.UpdateOne(
                {'_id': session_id},
                {'$set': {'roles_assigned': True}}
            )
        )

        # Execute bulk write if there are operations
        if bulk_operations:
            db.sessions.bulk_write(bulk_operations)

        # Start or restart the games
        for game_id in session.games:
            game = GAMES[game_id]
            game.start_game()  # Make sure this method properly resets and starts the game

        # Fetch updated session data
        updated_session_data = db.sessions.find_one({'_id': session_id})
        updated_players = updated_session_data.get('players', [])

        # Prepare the response data
        response_data = {
            'message': 'Roles and teams assigned and games started successfully',
            'assignments': assignments,
            'updated_players': updated_players
        }

        return jsonify(response_data), 200

    except Exception as e:
        logger.exception(f"An error occurred: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    
@app.route('/assign_random_roles/<session_id>', methods=['POST'])
@cross_origin()
def assign_random_roles(session_id):
    try:
        if session_id not in SESSIONS:
            return jsonify({'error': 'Session not found'}), 404

        session = SESSIONS[session_id]
        session_data = db.sessions.find_one({'_id': session_id})

        # Check if roles have already been assigned
        if session_data.get('roles_assigned', False):
            return jsonify({'error': 'Roles have already been assigned for this session'}), 400

        players = session_data.get('players', [])
        waiting_players = [p for p in players if p['role'] is None or p['team'] is None]

        # Get available roles and CPU roles from the first game
        first_game = GAMES[session.games[0]]
        all_roles = first_game.manual_stations_names + first_game.auto_stations_names
        cpu_roles = first_game.auto_stations_names

        # Prepare teams and roles
        teams = session.games
        roles_per_team = {team: all_roles.copy() for team in teams}

        # Randomly assign roles and teams
        import random
        random.shuffle(waiting_players)
        assignments = {}

        for player in waiting_players:
            # Find the team with the most available roles
            team = max(roles_per_team, key=lambda k: len(roles_per_team[k]))
            
            # Select a role, prioritizing non-CPU roles
            available_roles = [r for r in roles_per_team[team] if r not in cpu_roles]
            if not available_roles:
                available_roles = roles_per_team[team]
            
            role = random.choice(available_roles)
            
            assignments[player['uid']] = {'role': role, 'team': team}
            roles_per_team[team].remove(role)

            # If all roles for a team are assigned, remove the team
            if not roles_per_team[team]:
                del roles_per_team[team]

        # Assign remaining roles to CPU
        for team, roles in roles_per_team.items():
            for role in roles:
                assignments[f"CPU_{team}_{role}"] = {'role': role, 'team': team, 'is_cpu': True}

        # Update roles and teams in the database
        bulk_operations = []
        for player_uid, assignment in assignments.items():
            if not player_uid.startswith("CPU_"):
                bulk_operations.append(
                    pymongo.UpdateOne(
                        {'_id': session_id, 'players.uid': player_uid},
                        {'$set': {
                            'players.$.role': assignment['role'],
                            'players.$.team': assignment['team']
                        }}
                    )
                )

        # Add operation to set roles_assigned to True
        bulk_operations.append(
            pymongo.UpdateOne(
                {'_id': session_id},
                {'$set': {'roles_assigned': True}}
            )
        )

        # Execute bulk write
        if bulk_operations:
            db.sessions.bulk_write(bulk_operations)

        # Start the game for each team
        for game_id in session.games:
            game = GAMES[game_id]
            game.start_game()

        return jsonify({
            'message': 'Roles and teams randomly assigned and games started successfully',
            'assignments': assignments
        }), 200

    except Exception as e:
        logger.exception(f"An error occurred: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    
# Get session details
@app.route('/get_session/<session_id>', methods=['GET'])
@cross_origin()
def get_session(session_id):
    try:
        # First, try to get the session from memory
        if session_id in SESSIONS:
            session = SESSIONS[session_id]
            session_data = session.to_dict()
        else:
            # If not in memory, try to fetch from MongoDB
            session_data = db.sessions.find_one({'_id': session_id})
        
        if not session_data:
            return jsonify({'error': 'Session not found'}), 404

        # Include the roles_assigned variable in the response
        roles_assigned = session_data.get('roles_assigned', False)

        # Create a response dictionary with all session data and roles_assigned
        response_data = {
            **session_data,
            'roles_assigned': roles_assigned
        }

        return jsonify(response_data), 200

    except pymongo.errors.PyMongoError as e:
        logger.exception(f"A MongoDB error occurred: {str(e)}")
        return jsonify({'error': f'A database error occurred: {str(e)}'}), 500
    except Exception as e:
        logger.exception(f"An error occurred: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    

def monitor_games(session_id):
    try:
        session = SESSIONS.get(session_id)
        if not session:
            return jsonify({'error': 'Session not found'}), 404

        games_data = []
        for game_id in session.games:
            game = GAMES.get(game_id)
            if not game:
                continue  # Skip if game not found

            # Calculate game status
            if game.game_done:
                status = "Finished"
            elif game.current_week == 0:
                status = "Not Started"
            else:
                status = "Active"

            # Calculate players and orders
            total_players = len(game.manual_stations_names)
            active_players = sum(1 for station in game.manual_stations_names if game.network_stations[station].player_name)
            total_orders = len(game.manual_stations_names)
            completed_orders = sum(1 for station in game.manual_stations_names if game.network_stations[station].week_turn_completed >= game.current_week)

            # Calculate total cost
            total_cost = sum(game.kpi_cost['total']) if game.kpi_cost['total'] else 0

            # Calculate timer status (assuming turn_time is in seconds)
            current_time = datetime.datetime.now()
            if isinstance(game.turn_start_time, datetime.datetime):
                time_elapsed = (current_time - game.turn_start_time).total_seconds()
            elif isinstance(game.turn_start_time, (int, float)):
                time_elapsed = current_time.timestamp() - game.turn_start_time
            else:
                time_elapsed = 0
            
            timer_status = f"{int(time_elapsed)}/{game.turn_time}" if game.turn_time else "N/A"

            games_data.append({
                'team_name': game.team_name,
                'status': status,
                'players': f"{active_players}/{total_players}",
                'orders': f"{completed_orders}/{total_orders}",
                'week': f"{game.current_week}/{game.weeks}",
                'cost': f"$ {total_cost:,.0f}",
                'timer': timer_status
            })

        return jsonify({
            'session_name': session.name,
            'games': games_data
        }), 200

    except Exception as e:
        logger.exception(f"An error occurred while monitoring games: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    
# Route to access the monitor function
@app.route('/monitor_session/<session_id>', methods=['GET'])
@cross_origin()
def get_session_monitor(session_id):
    return monitor_games(session_id)

def load_sessions():
    logger.info("Loading sessions from MongoDB")
    try:
        sessions = db.sessions.find()
        for session_data in sessions:
            try:
                session = Session(
                    session_data['name'], 
                    session_data['owner_id'], 
                    session_data['num_teams']
                )
                session.id = str(session_data['_id'])  # MongoDB uses _id as the default identifier
                session.games = session_data['games']
                session.players = set(session_data['players'])
                SESSIONS[session.id] = session
                
                logger.info(f"Loaded session {session.id} with {len(session.games)} associated games")
            except Exception as e:
                logger.exception(f"Error loading session {session_data['_id']}: {str(e)}")

        logger.info(f"Loaded {len(SESSIONS)} sessions")
    except pymongo.errors.PyMongoError as e:
        logger.exception(f"A MongoDB error occurred while loading sessions: {str(e)}")
    except Exception as e:
        logger.exception(f"An unexpected error occurred while loading sessions: {str(e)}")

@app.route('/session/lobby/<session_id>', methods=['GET'])
@cross_origin()
def get_lobby_roles_status(session_id):
    try:
        if session_id not in SESSIONS:
            logger.error(f"Session {session_id} not found")
            return jsonify({'error': 'Session not found'}), 404
        
        session = SESSIONS[session_id]
        roles_status = {}
        current_time = time.time()

        for team_name in session.games:
            if team_name not in GAMES:
                logger.warning(f"Game {team_name} not found in session {session_id}")
                continue

            game_instance = GAMES[team_name]
            team_name = game_instance.team_name
            roles_status[team_name] = []

            for role, station in game_instance.network_stations.items():
                if isinstance(station, Demand):
                    # Skip Demand stations as they are not playable roles
                    continue

                is_cpu = station.auto_decide_order_qty
                is_connected = (current_time - station.last_communication_time) < SecondsAway_to_Disconnect
                is_player_name_set = station.player_name and station.player_name != f"player{station.station_name}"
                
                is_taken = is_cpu or (is_connected and is_player_name_set)

                player_status = 'CPU' if is_cpu else (
                    station.player_name if (is_connected and is_player_name_set) else (
                        'Disconnected' if is_player_name_set else 'Not Taken'
                    )
                )

                status = {
                    'role': role,
                    'taken': is_taken,
                    'player': player_status
                }
                roles_status[team_name].append(status)

        return jsonify(roles_status), 200

    except Exception as e:
        logger.exception(f"Error occurred while getting lobby roles status for session {session_id}: {str(e)}")
        return jsonify({'error': f"An error occurred: {str(e)}"}), 500
    
    
import copy

@app.route('/change_session_game_settings', methods=['POST'])
@cross_origin()
def change_session_game_settings():
    try:
        session_id = request.args.get('session')
        logger.info(f"Received request to change settings for session: {session_id}")
        
        if session_id not in SESSIONS:
            return jsonify({'error': 'Session not found'}), 404
        
        session = SESSIONS[session_id]
        game_settings = request.json
        
        # Process the game settings
        for x in game_settings.get('demands', []):
            x['demand'] = txt2list(x['demand'])
        
        for x in game_settings.get('stations', []):
            x['order_max'] = txt2list(x['order_max'])
            x['order_min'] = txt2list(x['order_min'])
            x['ship_max'] = txt2list(x['ship_max'])
            x['ship_min'] = txt2list(x['ship_min'])
        
        for x in game_settings.get('script', []):
            x['week'] = int(x['week'])
        
        updated_games = []
        
        # Apply settings to all games in the session
        for game_id in session.games:
            if game_id not in GAMES:
                logger.warning(f"Game {game_id} not found in GAMES dictionary")
                continue
            
            try:
                # Create a deep copy of game settings for this specific game
                game_specific_settings = copy.deepcopy(game_settings)
                game_specific_settings['team_name'] = GAMES[game_id].team_name
                game_specific_settings['_id'] = game_id  # Use _id for MongoDB
                
                # Create a new game object with the updated settings
                new_game = game.Game(game_specific_settings)
                
                # Update the game in the GAMES dictionary
                GAMES[game_id] = new_game
                
                # Update the game in MongoDB
                result = db.games.update_one(
                    {'_id': game_id},
                    {'$set': new_game.get_config()},
                    upsert=True
                )
                
                if result.modified_count > 0 or result.upserted_id:
                    updated_games.append(game_id)
                    logger.info(f"Updated game: {game_id}")
                else:
                    logger.warning(f"Game {game_id} was not updated in MongoDB")
                
            except pymongo.errors.PyMongoError as e:
                logger.error(f"MongoDB error updating game {game_id}: {str(e)}")
            except Exception as e:
                logger.error(f"Error updating game {game_id}: {str(e)}")
        
        logger.info(f"Successfully updated {len(updated_games)} out of {len(session.games)} games in session {session_id}")
        
        return jsonify({
            'result': True,
            'msg': f'Game setup saved for {len(updated_games)} games in the session.',
            'updated_games': updated_games
        })
    
    except pymongo.errors.PyMongoError as e:
        logger.error(f"MongoDB error occurred while saving game setup for session {session_id}: {str(e)}", exc_info=True)
        return jsonify({
            'result': False,
            'msg': f"Couldn't save game setup. Database error: {str(e)}."
        }), 500
    except Exception as e:
        logger.error(f"Error occurred while saving game setup for session {session_id}: {str(e)}", exc_info=True)
        return jsonify({
            'result': False,
            'msg': f"Couldn't save game setup. Error: {str(e)}."
        }), 500
    
@app.route('/get_session_game_settings', methods=['GET'])
@cross_origin()
def get_session_game_settings():
    try:
        session_id = request.args.get('session')
        logger.info(f"Received request to get settings for session: {session_id}")
        
        if session_id not in SESSIONS:
            return jsonify({'error': 'Session not found'}), 404
        
        session = SESSIONS[session_id]
        
        if not session.games:
            return jsonify({'error': 'No games found in the session'}), 404
        
        # Get the first game in the session
        first_game_id = session.games[0]
        if first_game_id not in GAMES:
            return jsonify({'error': 'Game not found'}), 404
        
        game_instance = GAMES[first_game_id]
        game_config = game_instance.get_config()
        
        # Process the game config
        game_config_processed = infinity_to_string(game_config)
        
        response = {
            'setup': game_config_processed,
            'action': 'edit'  # We use 'edit' to maintain consistency with edit_game_setup
        }
        
        logger.info(f"Successfully retrieved settings for session {session_id}")
        
        return jsonify(response), 200
    
    except Exception as e:
        logger.error(f"Error occurred while getting game settings for session {session_id}: {str(e)}", exc_info=True)
        return jsonify({
            'error': f"Couldn't get game settings. Error: {str(e)}."
        }), 500
    
@app.route('/get_session_analysis/<session_id>', methods=['GET'])
@cross_origin()
def get_session_analysis(session_id):
    if session_id not in SESSIONS:
        return jsonify({'error': 'Session not found'}), 404
    
    session = SESSIONS[session_id]
    analysis_data = {}
    
    for game_id in session.games:
        if game_id not in GAMES:
            continue  # Skip if game not found
        
        game = GAMES[game_id]
        analysis_data[game_id] = {
            'supply_chain_evolution': get_supply_chain_evolution(game),
            'stages_evolution': get_stages_evolution(game),
            'orders_vs_final_demand': get_orders_vs_final_demand(game),
            'stock_vs_final_demand': get_stock_vs_final_demand(game)
        }
    
    return jsonify(analysis_data)

# The helper functions (get_supply_chain_evolution, get_stages_evolution, etc.) remain the same
def get_supply_chain_evolution(game):
    weeks = range(1, game.current_week + 1)
    cost_data = {station: game.network_stations[station].kpi_total_cost[:game.current_week] for station in game.manual_stations_names + game.auto_stations_names}
    fill_rate_data = {station: game.network_stations[station].kpi_fulfilment_rate[:game.current_week] for station in game.manual_stations_names + game.auto_stations_names}
    avg_stock_data = {station: game.network_stations[station].inventory[:game.current_week] for station in game.manual_stations_names + game.auto_stations_names}
    
    return {
        'weeks': list(weeks),
        'cost': cost_data,
        'fill_rate': fill_rate_data,
        'average_stock': avg_stock_data
    }

def get_stages_evolution(game):
    weeks = range(1, game.current_week + 1)
    stages_data = {}
    for station in game.manual_stations_names + game.auto_stations_names:
        station_data = game.network_stations[station]
        stages_data[station] = {
            'stock': station_data.inventory[:game.current_week],
            'backorder': [sum(station_data.backorder[customer][week] for customer in station_data.backorder) for week in range(game.current_week)],
            'order': [sum(station_data.sent_po[supplier][week] for supplier in station_data.sent_po) for week in range(game.current_week)],
            'cost': station_data.kpi_total_cost[:game.current_week]
        }
    
    return {
        'weeks': list(weeks),
        'stages': stages_data
    }

def get_orders_vs_final_demand(game):
    weeks = range(1, game.current_week + 1)
    orders_data = {station: [sum(game.network_stations[station].sent_po[supplier][week] for supplier in game.network_stations[station].sent_po) for week in range(game.current_week)] for station in game.manual_stations_names + game.auto_stations_names}
    final_demand = game.network_stations[game.demand_stations_names[0]].demand[:game.current_week]
    
    return {
        'weeks': list(weeks),
        'orders': orders_data,
        'final_demand': final_demand
    }

def get_stock_vs_final_demand(game):
    weeks = range(1, game.current_week + 1)
    stock_data = {station: game.network_stations[station].inventory[:game.current_week] for station in game.manual_stations_names + game.auto_stations_names}
    final_demand = game.network_stations[game.demand_stations_names[0]].demand[:game.current_week]
    
    return {
        'weeks': list(weeks),
        'stock': stock_data,
        'final_demand': final_demand
    }

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import os

@app.route('/generate_session_excel/<session_id>', methods=['GET'])
@cross_origin()
def generate_session_excel(session_id):
    try:
        session = SESSIONS.get(session_id)
        if not session:
            return jsonify({'error': 'Session not found'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Session Overview"

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Session Overview
        ws['A1'] = "Session Name"
        ws['B1'] = session.name
        ws['A2'] = "Number of Teams"
        ws['B2'] = session.num_teams

        # Teams Overview
        ws['A4'] = "Team Performance Overview"
        ws['A4'].font = header_font
        ws.merge_cells('A4:G4')

        headers = ["Team", "Total Cost", "Avg. Fulfillment Rate", "Avg. Green Score", "Current Week"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 6
        for game_id in session.games:
            game = GAMES.get(game_id)
            if game:
                ws.cell(row=row, column=1, value=game.team_name)
                ws.cell(row=row, column=2, value=sum(game.kpi_cost['total']))
                ws.cell(row=row, column=3, value=sum(game.kpi_customer_satisfaction) / len(game.kpi_customer_satisfaction))
                ws.cell(row=row, column=4, value=sum(game.kpi_green_score) / len(game.kpi_green_score))
                ws.cell(row=row, column=5, value=game.current_week)
                row += 1

        # Individual Player Performance
        for game_id in session.games:
            game = GAMES.get(game_id)
            if game:
                ws = wb.create_sheet(title=f"Team {game.team_name}")
                
                ws['A1'] = f"Team {game.team_name} - Player Performance"
                ws['A1'].font = header_font
                ws.merge_cells('A1:G1')

                headers = ["Player", "Station", "Total Cost", "Avg. Fulfillment Rate", "Avg. Green Score", "Inventory", "Backorders"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                row = 3
                for station_name, station in game.network_stations.items():
                    if not isinstance(station, Demand):  # Skip Demand stations
                        ws.cell(row=row, column=1, value=station.player_name)
                        ws.cell(row=row, column=2, value=station_name)
                        ws.cell(row=row, column=3, value=sum(station.kpi_total_cost))
                        ws.cell(row=row, column=4, value=sum(station.kpi_fulfilment_rate) / len(station.kpi_fulfilment_rate))
                        ws.cell(row=row, column=5, value=sum(station.kpi_truck_utilization) / len(station.kpi_truck_utilization))
                        ws.cell(row=row, column=6, value=station.inventory[-1] if station.inventory else 0)
                        ws.cell(row=row, column=7, value=sum(station.backorder[customer][-1] for customer in station.backorder))
                        row += 1

                # Weekly Performance
                ws['A10'] = "Weekly Performance"
                ws['A10'].font = header_font
                ws.merge_cells('A10:G10')

                headers = ["Week", "Cost", "Fulfillment Rate", "Green Score", "Shipments", "Inventory", "Backorders"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=11, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                row = 12
                for week in range(game.current_week):
                    ws.cell(row=row, column=1, value=week + 1)
                    ws.cell(row=row, column=2, value=sum(station.kpi_total_cost[week] for station in game.network_stations.values() if not isinstance(station, Demand)))
                    ws.cell(row=row, column=3, value=sum(station.kpi_fulfilment_rate[week] for station in game.network_stations.values() if not isinstance(station, Demand)) / len(game.network_stations))
                    ws.cell(row=row, column=4, value=sum(station.kpi_truck_utilization[week] for station in game.network_stations.values() if not isinstance(station, Demand)) / len(game.network_stations))
                    ws.cell(row=row, column=5, value=sum(station.kpi_shipment_trucks[week] for station in game.network_stations.values() if not isinstance(station, Demand)))
                    ws.cell(row=row, column=6, value=sum(station.inventory[week] for station in game.network_stations.values() if not isinstance(station, Demand)))
                    ws.cell(row=row, column=7, value=sum(sum(station.backorder[customer][week] for customer in station.backorder) for station in game.network_stations.values() if not isinstance(station, Demand)))
                    row += 1

        # Save the workbook
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'session_{session_id}_data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.exception(f"An error occurred while generating Excel for session {session_id}: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

load_sessions()